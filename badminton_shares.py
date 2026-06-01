"""
Court-Share Calculator
======================

Given a list of polls (each: title, Yes voters, payer), parse each title for
the number of courts and duration, compute the total cost at a fixed rate
per court per hour, and split it equally among the Yes voters. Writes an
.xlsx report with four sheets: Polls, Owes <collector>, Settlement,
Who-owes-whom.

This module is consumed both by the standalone CLI (run directly to compute
a one-off report) and by `fetch_telegram_polls.py` (the Telegram bot uses
`compute_cost`, `build_collector_view`, and `write_report` from here).

CLI usage:
    pip3 install openpyxl --break-system-packages
    python3 badminton_shares.py --input polls.json --output month.xlsx
    python3 badminton_shares.py --collector "Sara Lee"      # who collects

If --input is omitted, the hardcoded POLLS list at the bottom is used. The
JSON input format is whatever `fetch_telegram_polls.py` writes, namely a
list of records each shaped like:
    { "date": "2026-05-31", "title": "...", "sender": "Milad",
      "options": [{"name": "Yes", "voters": ["Milad", ...]},
                  {"name": "No",  "voters": [...]}] }
The first option's voters list is treated as the Yes side and the poll's
"sender" as the payer.
"""

import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Default court rate
RATE_PER_COURT_PER_HOUR = 25.0


@dataclass
class Poll:
    """One booking session, derived from a poll.

    Fields:
      date:        ISO date, e.g. "2026-05-05".
      title:       raw poll title — the title parser pulls courts/hours out.
      yes_voters:  list of display names who voted Yes. Each pays an equal share.
      payer:       who fronted the court fee. Others owe them their share.
      override_*:  short-circuit the title parser when you want to set them
                   explicitly (e.g., the title is ambiguous or wrong).
    """

    date: str
    title: str
    yes_voters: List[str]
    payer: str = "Milad"
    override_amount: Optional[float] = None
    override_courts: Optional[int] = None
    override_hours: Optional[float] = None


# --- Title parsing -----------------------------------------------------------

def parse_courts(title: str) -> int:
    """
    Return the number of courts implied by the title.

    Conventions seen in the chat:
      "(Court 3)"          -> 1 court (court #3)
      "(Court 2,3)"        -> 2 courts (courts #2 and #3)
      "(Court 3,4)"        -> 2 courts
      "(court2)"           -> 1 court (court #2, no space)
    """
    # Look for the parenthesized court spec
    m = re.search(r"\(\s*[Cc]ourt[s]?\s*([0-9,\s]+)\)", title)
    if not m:
        return 1  # default to 1 court if title doesn't specify
    spec = m.group(1).strip()
    # Split on commas to count distinct court numbers
    parts = [p.strip() for p in spec.split(",") if p.strip()]
    return max(1, len(parts))


def parse_hours(title: str) -> float:
    """
    Return the booking duration in hours, parsed from the title.

    Handles:
      "5:00-6:00 pm"
      "8:00-9:00 pm"
      "12:00-1:00 pm"
      "8 PM"   -> assume 1 hour
    """
    # Require a "H:MM-H:MM" range. Both sides must have :MM so we don't match
    # things like "May26 - 8 PM" (a single time, not a range).
    m = re.search(
        r"(\d{1,2}):(\d{2})\s*[-–]\s*(\d{1,2}):(\d{2})\s*(am|pm)?",
        title,
        flags=re.IGNORECASE,
    )
    if m:
        h1 = int(m.group(1)); m1 = int(m.group(2))
        h2 = int(m.group(3)); m2 = int(m.group(4))
        # crude AM/PM wrap fix: badminton sessions are short, so if end<start
        # we assume they cross noon (e.g., 12:00 -> 1:00).
        start = h1 * 60 + m1
        end = h2 * 60 + m2
        if end <= start:
            end += 12 * 60
        return (end - start) / 60.0
    # No range -> default 1 hour
    return 1.0


def compute_cost(poll: Poll) -> tuple[int, float, float]:
    """Return (courts, hours, total_cost) for a poll."""
    courts = poll.override_courts or parse_courts(poll.title)
    hours = poll.override_hours or parse_hours(poll.title)
    if poll.override_amount is not None:
        cost = poll.override_amount
    else:
        cost = courts * hours * RATE_PER_COURT_PER_HOUR
    return courts, hours, cost


# --- Spreadsheet output -------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD = Font(bold=True)


def autosize(ws):
    for col_cells in ws.columns:
        col = get_column_letter(col_cells[0].column)
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col].width = min(max(width + 2, 10), 40)


def build_collector_view(polls: List[Poll], collector: str):
    """
    Settlement from one person's perspective ("everyone pays {collector}, then
    {collector} reimburses anyone who fronted cash").

    For each other person:
        net_due_to_collector = their_total_share - their_total_paid_out
        > 0  → they pay collector that much
        < 0  → collector pays them that much

    Returns (rows, totals) where:
        rows = [(person, action, amount), ...] sorted high → low by amount
        totals = dict with "inflow", "outflow", "net"
    """
    person_share = defaultdict(float)
    person_paid = defaultdict(float)
    for poll in polls:
        _, _, cost = compute_cost(poll)
        n = len(poll.yes_voters)
        if not n:
            continue
        share = cost / n
        for v in poll.yes_voters:
            person_share[v] += share
        person_paid[poll.payer] += cost

    rows = []
    inflow = outflow = 0.0
    people = sorted(set(list(person_share.keys()) + list(person_paid.keys())))
    for p in people:
        if p == collector:
            continue
        net = person_share[p] - person_paid[p]
        if abs(net) < 0.01:
            continue
        if net > 0:
            rows.append((p, f"Pays {collector}", net))
            inflow += net
        else:
            rows.append((p, f"{collector} pays them", -net))
            outflow += -net
    rows.sort(key=lambda r: -r[2])
    return rows, {"inflow": inflow, "outflow": outflow, "net": inflow - outflow}


def write_report(polls: List[Poll], output_path: str, collector: str = "Milad"):
    wb = Workbook()

    # --- Sheet 1: Per-poll breakdown ----------------------------------------
    ws1 = wb.active
    ws1.title = "Polls"
    headers = ["Date", "Title", "Courts", "Hours", "Total Cost", "Yes Voters", "# Yes", "Share / Person", "Payer"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    total_cost_sum = 0.0
    for i, poll in enumerate(polls, start=2):
        courts, hours, cost = compute_cost(poll)
        n = len(poll.yes_voters)
        share = cost / n if n else 0.0
        total_cost_sum += cost

        ws1.cell(row=i, column=1, value=poll.date)
        ws1.cell(row=i, column=2, value=poll.title)
        ws1.cell(row=i, column=3, value=courts)
        ws1.cell(row=i, column=4, value=hours)
        ws1.cell(row=i, column=5, value=cost).number_format = "$#,##0.00"
        ws1.cell(row=i, column=6, value=", ".join(poll.yes_voters))
        ws1.cell(row=i, column=7, value=n)
        ws1.cell(row=i, column=8, value=share).number_format = "$#,##0.00"
        ws1.cell(row=i, column=9, value=poll.payer)

        if i % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=i, column=c).fill = ALT_FILL

    last = len(polls) + 1
    ws1.cell(row=last + 1, column=4, value="TOTAL").font = BOLD
    ws1.cell(row=last + 1, column=5, value=total_cost_sum).number_format = "$#,##0.00"
    ws1.cell(row=last + 1, column=5).font = BOLD
    autosize(ws1)

    # --- Sheet 2: Who owes whom ---------------------------------------------
    ws2 = wb.create_sheet("Settlement")
    headers2 = ["Person", "Sessions Played (Yes)", "Total Owed", "Status"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Accumulate per-person debt grouped by payer (assume one payer for now;
    # if mixed, "Total Owed" is what they owe across all polls combined,
    # netted against what they paid).
    person_share = defaultdict(float)   # what each person owes for sessions
    person_paid = defaultdict(float)    # what each person paid out of pocket
    person_sessions = defaultdict(int)

    for poll in polls:
        courts, hours, cost = compute_cost(poll)
        n = len(poll.yes_voters)
        if not n:
            continue
        share = cost / n
        for v in poll.yes_voters:
            person_share[v] += share
            person_sessions[v] += 1
        person_paid[poll.payer] += cost

    # Net = paid - owed.  Positive net = owed money back; negative = owes.
    all_people = sorted(set(list(person_share.keys()) + list(person_paid.keys())))

    row = 2
    for p in all_people:
        net = person_paid[p] - person_share[p]
        if net > 0.01:
            status = f"Is owed ${net:,.2f}"
        elif net < -0.01:
            status = f"Owes ${-net:,.2f}"
        else:
            status = "Settled"
        ws2.cell(row=row, column=1, value=p)
        ws2.cell(row=row, column=2, value=person_sessions[p])
        ws2.cell(row=row, column=3, value=person_share[p]).number_format = "$#,##0.00"
        ws2.cell(row=row, column=4, value=status)
        if row % 2 == 0:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    autosize(ws2)

    # --- Sheet 3: Who-owes-whom (pairwise, per payer) -----------------------
    ws3 = wb.create_sheet("Who owes whom")
    headers3 = ["Debtor", "Creditor (payer)", "Amount"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    pair_debt = defaultdict(float)  # (debtor, creditor) -> amount
    for poll in polls:
        courts, hours, cost = compute_cost(poll)
        n = len(poll.yes_voters)
        if not n:
            continue
        share = cost / n
        for v in poll.yes_voters:
            if v != poll.payer:
                pair_debt[(v, poll.payer)] += share

    row = 2
    for (debtor, creditor), amount in sorted(pair_debt.items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=debtor)
        ws3.cell(row=row, column=2, value=creditor)
        ws3.cell(row=row, column=3, value=amount).number_format = "$#,##0.00"
        if row % 2 == 0:
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=c).fill = ALT_FILL
        row += 1

    autosize(ws3)

    # --- Sheet 4: Owes Me (collector-centric) -------------------------------
    ws4 = wb.create_sheet(f"Owes {collector}")
    # Put it second so it's the first thing after the raw poll list.
    wb.move_sheet(ws4, offset=-2)

    title_cell = ws4.cell(row=1, column=1,
                          value=f'"Everyone pays {collector}" settlement')
    title_cell.font = Font(bold=True, size=14)
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    headers4 = ["Person", "Action", "Amount"]
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    rows, totals = build_collector_view(polls, collector)
    r = 4
    for person, action, amount in rows:
        ws4.cell(row=r, column=1, value=person)
        ws4.cell(row=r, column=2, value=action)
        ws4.cell(row=r, column=3, value=amount).number_format = "$#,##0.00"
        if r % 2 == 0:
            for c in range(1, len(headers4) + 1):
                ws4.cell(row=r, column=c).fill = ALT_FILL
        r += 1

    # Totals block
    r += 1
    ws4.cell(row=r, column=1, value="Total collected").font = BOLD
    ws4.cell(row=r, column=3, value=totals["inflow"]).number_format = "$#,##0.00"
    ws4.cell(row=r, column=3).font = BOLD
    r += 1
    ws4.cell(row=r, column=1, value=f"Total {collector} pays out").font = BOLD
    ws4.cell(row=r, column=3, value=totals["outflow"]).number_format = "$#,##0.00"
    ws4.cell(row=r, column=3).font = BOLD
    r += 1
    ws4.cell(row=r, column=1, value=f"{collector}'s net").font = BOLD
    ws4.cell(row=r, column=3, value=totals["net"]).number_format = "$#,##0.00"
    ws4.cell(row=r, column=3).font = BOLD
    for c in range(1, len(headers4) + 1):
        ws4.cell(row=r, column=c).fill = SUBTOTAL_FILL

    autosize(ws4)

    wb.save(output_path)
    print(f"Wrote {output_path}")


# --- Poll data (May 2026) ----------------------------------------------------
# Edit this list to add/update polls. The script will recompute the xlsx.

POLLS: List[Poll] = [
    Poll(
        date="2026-05-05",
        title="Tuesday May 5th (Court 3,4) - 5:00-6:00 pm",
        yes_voters=["Milad", "Mohammad Ahmadi", "Thoufiq Serajaldin",
                    "Kapeesh Kaul", "Rahim Samei", "Sathyajit Loganathan"],
        payer="Kapeesh Kaul",   # Kapeesh covered this one
    ),
    Poll(
        date="2026-05-08",
        title="Friday May 8th (Court 4) - 8:00-9:00 pm",
        yes_voters=["Milad", "Rahim Samei", "Elnaz Yousefi", "Mohammad Ahmadi"],
    ),
    Poll(
        date="2026-05-10",
        title="Sunday May 10th (Court 2,3) - 12:00-1:00 pm",
        yes_voters=["Milad", "Rahim Samei", "Nick"],
    ),
    Poll(
        date="2026-05-12",
        title="Tuesday May 12th (Court 4) - 8:00-9:00 pm",
        yes_voters=["Milad", "Mohammad Ahmadi"],
    ),
    Poll(
        date="2026-05-16",
        title="Saturday May 16th (Court 2) - 5:00-6:00 pm",
        yes_voters=["Milad", "Mohammad Ahmadi"],
    ),
    Poll(
        date="2026-05-17",
        title="Sunday May 17th (Court 3) - 12:00-1:00 pm",
        yes_voters=["Milad", "Sathyajit Loganathan"],
    ),
    Poll(
        date="2026-05-23",
        title="Saturday May 23rd (Court 2) - 5:00-6:00 pm",
        yes_voters=["Milad"],   # only 1 Yes vote -- session likely cancelled
    ),
    Poll(
        date="2026-05-24",
        title="Sunday May 24th (Court 2) - 12:00-1:00 pm",
        yes_voters=["Milad"],   # only 1 Yes vote -- session likely cancelled
    ),
    Poll(
        date="2026-05-26",
        title="Tuesday May26 - 8 PM (court2)",
        yes_voters=["Milad", "Mohammad Ahmadi"],
    ),
    Poll(
        date="2026-05-30",
        title="Saturday May 30th (Court 3) - 5:00-6:00 pm",
        yes_voters=["Milad", "Rahim Samei", "Thoufiq Serajaldin", "Kapeesh Kaul"],
    ),
    Poll(
        date="2026-05-31",
        title="Sunday May 31st (Court 3) - 12:00-1:00 pm",
        yes_voters=["Milad", "Mohammad Ahmadi", "Elnaz Yousefi",
                    "Rahim Samei", "Thoufiq Serajaldin"],
    ),
]


def load_polls_from_json(path: str) -> List[Poll]:
    """
    Load polls from a JSON file produced by fetch_polls.js.

    Expected shape (per poll):
      { "id": "...", "date": "YYYY-MM-DD", "title": "...",
        "sender": "Milad",
        "options": [ {"name": "Yes", "voters": ["Milad", ...]},
                     {"name": "No",  "voters": ["..."]} ] }

    The FIRST option is treated as the "Yes" option (matches how the badminton
    polls are written). The poll creator (sender) is treated as the payer.
    """
    import json
    with open(path) as f:
        raw = json.load(f)
    polls: List[Poll] = []
    for p in raw:
        opts = p.get("options") or []
        yes_voters = opts[0].get("voters", []) if opts else []
        polls.append(
            Poll(
                date=p["date"],
                title=p.get("title", ""),
                yes_voters=list(yes_voters),
                payer=p.get("sender") or "Milad",
            )
        )
    # Sort by date so the spreadsheet reads chronologically.
    polls.sort(key=lambda x: x.date)
    return polls


if __name__ == "__main__":
    import argparse, os
    here = os.path.dirname(os.path.abspath(__file__))
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input", "-i",
        help="Path to polls.json (output of fetch_polls.js). "
             "If omitted, uses the hardcoded POLLS list in this file.",
    )
    parser.add_argument(
        "--output", "-o",
        default=os.path.join(here, "badminton_shares.xlsx"),
        help="Path to write the .xlsx report.",
    )
    parser.add_argument(
        "--collector", "-c",
        default="Milad",
        help='Who is collecting from everyone (default "Milad"). '
             'Drives the "Owes <collector>" sheet.',
    )
    args = parser.parse_args()

    polls = load_polls_from_json(args.input) if args.input else POLLS
    write_report(polls, args.output, collector=args.collector)
